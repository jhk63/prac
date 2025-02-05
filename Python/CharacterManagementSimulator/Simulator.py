
from openpyxl import load_workbook
from openpyxl import Workbook
import plotly.express as px


# load_wb = load_workbook("sample.xlsx", data_only=True)
# load_wb = load_workbook("Python/BattleSimulator/sample.xlsx", data_only=True)
# load_ws = load_wb.active

# print('\n-----모든 행 단위로 출력-----')
# row_value = []
# for row in load_ws.rows:
#     for cell in row:
#         row_value.append(cell.value)

# print(row_value)

# row_value = [row.value for row in load_ws[3]]
# print(row_value)
# print(type(row_value[1]))
# print(type(row_value[5]))

# print('\n-----지정한 셀 출력-----')
# get_cells = load_ws['A3':'I3']
# for row in get_cells:
#     for cell in row:
        # print(cell.value)


class Character:
    def __init__(self, name, hp, defense, attack, speed, crit_rate, crit_damage, accuracy, dodge):
        self.name = name
        self.hp = hp
        self.defense = defense
        self.attack = attack
        self.speed = speed
        self.crit_rate = crit_rate
        self.crit_damage = crit_damage
        self.accuracy = accuracy
        self.dodge = dodge
        self.combat_power = self.calculate_combat_power()

    def calculate_combat_power(self):
        return (self.hp * 0.2) + \
                (self.defense * 2) + \
                (self.attack * 3) + \
                (self.speed * 100) + \
                (self.crit_rate * self.crit_damage * 1.5) + \
                (self.accuracy * 10) + \
                (self.dodge * 15)


class Simulator:
    def __init__(self, filepath):
        self.filepath = filepath
        self.characters = self.load_characters()

    def load_characters(self):
        load_wb = load_workbook(self.filepath, data_only=True)
        load_ws = load_wb.active
        characters = []

        for row in load_ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                break

            char = Character(
                name=row[0],
                hp=float(row[1]),
                defense=float(row[2]),
                attack=float(row[3]),
                speed=float(row[4]),
                crit_rate=float(row[5]),
                crit_damage=float(row[6]),
                accuracy=float(row[7]),
                dodge=float(row[8])
            )
            characters.append(char)
        
        load_wb.close()
        return characters
    
    def add_character(self, name, hp, defense, attack, speed, crit_rate, crit_damage, accuracy, dodge):
        load_wb = load_workbook(self.filepath, data_only=True)
        load_ws = load_wb.active
        
        new_row = [
            name, hp, defense, attack, speed, 
            crit_rate, crit_damage, accuracy, dodge
        ]

        load_ws.append(new_row)
        load_wb.save(self.filepath)
        load_wb.close()
        
        char = Character(name, hp, defense, attack, speed, crit_rate, crit_damage, accuracy, dodge)
        self.characters.append(char)

    def delete_character(self, name):
        load_wb = load_workbook(self.filepath, data_only=True)
        load_ws = load_wb.active
        
        for idx, row in enumerate(load_ws.iter_rows(min_row=3, values_only=True), start=3):
            if row[0] == name:
                load_ws.delete_rows(idx)
                break
        
        load_wb.save(self.filepath)
        load_wb.close()
        
        self.characters = [char for char in self.characters if char.name != name]

    def print_characters(self):
        for char in self.characters:
            print(f"Name: {char.name}, HP: {char.hp}, Power: {char.combat_power}")

    def win_probability(self, attacker, defender):
        return attacker.calculate_combat_power() / (attacker.calculate_combat_power() + defender.calculate_combat_power())
    
    def simulate_all_battles(self):
        results = []
        for char1 in self.characters:
            char_results = []
            for char2 in self.characters:
                if char1 != char2:
                    win_prob = self.win_probability(char1, char2)
                    char_results.append((char2.name, win_prob))
            results.append((char1, char_results))
        return results

    def save_results_excel(self, output_filepath):
        save_wb = Workbook()
        save_ws = save_wb.active
        save_ws.title = "Character Stats"

        header = ["Name", "Combat Power", "Hp", "Defense", "Attack", "Speed", "Critical Rate", "Critical Damage", "Accuracy", "Dodge"]
        save_ws.append(header)

        for char in self.characters:
            row = [char.name, char.combat_power, char.hp, char.defense, char.attack, char.speed, char.crit_rate, char.crit_damage, char.accuracy, char.dodge]
            save_ws.append(row)
        
        win_ws = save_wb.create_sheet(title="Character Win Probabilities")
        header = ["Character"] + [char.name for char in self.characters]
        win_ws.append(header)

        for char1 in self.characters:
            row = [char1.name]
            for char2 in self.characters:
                win_prob = self.win_probability(char1, char2) * 100
                row.append(win_prob)
            win_ws.append(row)

        save_wb.save(output_filepath)
        print(f"\nSimulation results saved to '{output_filepath}'")

    def visualize_combat_power(self):
        names = [char.name for char in self.characters]
        powers = [char.combat_power for char in self.characters]

        fig = px.bar(x=names, y=powers, title="Character Combat Power",
                 labels={'x': 'Character', 'y': 'Combat Power'})
        return fig

    def visualize_win_probabilities(self):
        characters = [char.name for char in self.characters]
        num_chars = len(characters)
        
        # 히트맵 데이터를 저장할 2D 배열 생성
        win_prob_matrix = [[0] * num_chars for _ in range(num_chars)]
        
        # 데이터를 2D 배열에 채우기
        for i, char1 in enumerate(self.characters):
            for j, char2 in enumerate(self.characters):
                win_prob = self.win_probability(char1, char2) * 100
                win_prob_matrix[i][j] = win_prob

        custom_colorscale = [
            [0.0, "blue"],  # 0%
            [0.3, "lightblue"],  # 30% 
            [0.5, "yellow"],  # 50% 
            [0.7, "orange"],  # 70% 
            [1.0, "red"]  # 100% 
        ]

        # Plotly 히트맵 생성
        fig = px.imshow(
            win_prob_matrix,
            x=characters,
            y=characters,
            color_continuous_scale=custom_colorscale,
            range_color=[30, 60],
            labels={"x": "Opponent", "y": "Player", "color": "Win Probability (%)"},
            title="Character Win Probabilities Heatmap"
        )
        
        return fig


if __name__ == "__main__":    
    simulator = Simulator("Python/BattleSimulator/sample.xlsx")

    print("\n=== 전투력 출력 ===")
    simulator.print_characters()

    print("\n=== 전투 시뮬레이션 결과 ===")
    results = simulator.simulate_all_battles()
    for char, battles in results:
        print(f"\n{char.name} (전투력: {int(char.combat_power)})")
        for opponent_name, win_prob in battles:
            print(f"vs {opponent_name}: 승리확률 {win_prob}%")

    # simulator.visualize_combat_power().show()
    simulator.visualize_win_probabilities().show()

    simulator.save_results_excel("Python/BattleSimulator/output.xlsx")
